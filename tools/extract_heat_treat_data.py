import json
import math
import pathlib
from datetime import datetime
from typing import Any

import openpyxl

ROOT = pathlib.Path(__file__).resolve().parents[1]
EXCEL_PATH = ROOT / 'Heat Treat Verification Measurements.xlsm'
OUTPUT_DIR = ROOT / 'Server' / 'wwwroot' / 'data'
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

def coerce_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, float):
        if math.isfinite(value):
            return round(value, 6)
        return None
    return value

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

master_ws = wb['Master']
headers = [cell.value for cell in next(master_ws.iter_rows(min_row=1, max_row=1))]
records = []
for row in master_ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    record = {}
    for header, value in zip(headers, row):
        if header is None:
            continue
        record[str(header)] = coerce_value(value)
    records.append(record)

(master_path := OUTPUT_DIR / 'heatTreatMaster.json').write_text(json.dumps(records, indent=2))

ref_ws = wb['List References']
ref_data = {
    'typeOfTestOptions': [],
    'shiftOptions': [],
    'lineOptions': [],
    'passFailOptions': [],
    'hardnessSpecs': [],
    'caseDepthRanges': [],
    'judgementOptions': [],
    'operatorIds': [],
    'operatorNames': []
}
for row in ref_ws.iter_rows(min_row=1, values_only=True):
    if all(value is None for value in row):
        continue
    type_of_test, shift, line, pass_fail, hardness, case_depth, judgement, operator_id, operator_name = (
        row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9]
    )
    if type_of_test and type_of_test not in ref_data['typeOfTestOptions']:
        ref_data['typeOfTestOptions'].append(type_of_test)
    if shift and shift not in ref_data['shiftOptions']:
        ref_data['shiftOptions'].append(shift)
    if line and line not in ref_data['lineOptions']:
        ref_data['lineOptions'].append(line)
    if pass_fail and pass_fail not in ref_data['passFailOptions']:
        ref_data['passFailOptions'].append(pass_fail)
    if hardness and hardness not in ref_data['hardnessSpecs']:
        ref_data['hardnessSpecs'].append(hardness)
    if case_depth and case_depth not in ref_data['caseDepthRanges']:
        ref_data['caseDepthRanges'].append(case_depth)
    if judgement and judgement not in ref_data['judgementOptions']:
        ref_data['judgementOptions'].append(judgement)
    if operator_id and operator_id not in ref_data['operatorIds']:
        ref_data['operatorIds'].append(str(operator_id))
    if operator_name and operator_name not in ref_data['operatorNames']:
        ref_data['operatorNames'].append(operator_name)

(ref_path := OUTPUT_DIR / 'heatTreatListReferences.json').write_text(json.dumps(ref_data, indent=2))

print(f"Wrote {len(records)} master records to {master_path}")
print(f"Wrote list reference data to {ref_path}")
